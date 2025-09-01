from flask import Flask, render_template, request, redirect, url_for, flash
import sqlite3
from datetime import datetime
import os
from contextlib import contextmanager
import sys
from flask import send_file, request
import io
import xlsxwriter
try:
    import openpyxl
except ImportError:
    openpyxl = None
    # Optionally, you can log or print a warning here

# Helper to get resource path (for PyInstaller compatibility)
def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    if getattr(sys, 'frozen', False):
        # Running in a bundle (PyInstaller)
        base_path = os.path.dirname(sys.executable)
    else:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# Update DATABASE path for PyInstaller
DATABASE = resource_path('shipping.db')

# Update Flask template and static folder paths for PyInstaller
if hasattr(sys, '_MEIPASS'):
    app = Flask(__name__,
                template_folder=os.path.join(sys._MEIPASS, 'templates'),
                static_folder=os.path.join(sys._MEIPASS, 'static'))
else:
    app = Flask(__name__,
                template_folder='templates',
                static_folder='static')

app.config['SECRET_KEY'] = 'your-secret-key-here'

def init_db():
    with get_db() as db:
        db.execute('''
            CREATE TABLE IF NOT EXISTS shipment (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                shopiny_number TEXT UNIQUE NOT NULL,
                receipt_number TEXT UNIQUE,
                order_number TEXT,
                delivery_date DATETIME,
                from_governorate TEXT NOT NULL,
                to_governorate TEXT NOT NULL,
                carrier_company TEXT,
                notes TEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        db.execute('''
            CREATE TABLE IF NOT EXISTS shipment_type (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL
            )
        ''')
        
        db.execute('''
            CREATE TABLE IF NOT EXISTS department (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL
            )
        ''')
        
        db.execute('''
            CREATE TABLE IF NOT EXISTS shipment_item (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                shipment_id INTEGER NOT NULL,
                shipment_type_id INTEGER NOT NULL,
                department_id INTEGER NOT NULL,
                quantity INTEGER NOT NULL,
                cost INTEGER NOT NULL,
                boxes_count INTEGER NOT NULL,
                total INTEGER NOT NULL,
                notes TEXT,
                FOREIGN KEY (shipment_id) REFERENCES shipment (id) ON DELETE CASCADE,
                FOREIGN KEY (shipment_type_id) REFERENCES shipment_type (id),
                FOREIGN KEY (department_id) REFERENCES department (id)
            )
        ''')
        
        db.execute('''
            CREATE TABLE IF NOT EXISTS carrier_company (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL
            )
        ''')
        
        db.execute('''
            CREATE TABLE IF NOT EXISTS governorate (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL
            )
        ''')
    
        # إضافة عمود use_boxes إذا لم يكن موجوداً
        try:
            db.execute('ALTER TABLE shipment_item ADD COLUMN use_boxes INTEGER DEFAULT 0')
        except Exception:
            pass  # العمود موجود مسبقاً أو حدث خطأ آخر

@contextmanager
def get_db():
    db = sqlite3.connect(DATABASE)
    db.row_factory = sqlite3.Row
    try:
        yield db
        db.commit()
    except Exception as e:
        db.rollback()
        raise e
    finally:
        db.close()

@app.route('/')
def index():
    from_governorate = request.args.get('from_governorate')
    to_governorate = request.args.get('to_governorate')
    carrier_company = request.args.get('carrier_company')
    filter_field = request.args.get('filter_field')
    filter_value = request.args.get('filter_value')
    with get_db() as db:
        db.row_factory = sqlite3.Row
        query = '''
            SELECT s.*, GROUP_CONCAT(si.total) as total_amount,
                   datetime(s.delivery_date) as delivery_date
            FROM shipment s
            LEFT JOIN shipment_item si ON s.id = si.shipment_id
            WHERE 1=1
        '''
        params = []
        if from_governorate:
            query += ' AND s.from_governorate = ?'
            params.append(from_governorate)
        if to_governorate:
            query += ' AND s.to_governorate = ?'
            params.append(to_governorate)
        if carrier_company:
            query += ' AND s.carrier_company = ?'
            params.append(carrier_company)
        if filter_field in ['shopiny_number', 'receipt_number', 'order_number'] and filter_value:
            query += f' AND s.{filter_field} LIKE ?'
            params.append(f'%{filter_value}%')
        query += '\nGROUP BY s.id\nORDER BY s.created_at DESC'
        shipments = db.execute(query, params).fetchall()
        governorates = db.execute('SELECT name FROM governorate').fetchall()
        carrier_companies = db.execute('SELECT name FROM carrier_company').fetchall()
        processed_shipments = []
        for shipment in shipments:
            shipment_dict = dict(shipment)
            try:
                shipment_dict['delivery_date'] = datetime.strptime(shipment['delivery_date'], '%Y-%m-%d %H:%M:%S')
            except (ValueError, TypeError):
                shipment_dict['delivery_date'] = None
            processed_shipments.append(shipment_dict)
    return render_template(
        'index.html',
        shipments=processed_shipments,
        governorates=governorates,
        carrier_companies=carrier_companies,
        selected_from_governorate=from_governorate,
        selected_to_governorate=to_governorate,
        selected_carrier_company=carrier_company,
        filter_field=filter_field,
        filter_value=filter_value
    )

@app.route('/shipment/new', methods=['GET', 'POST'])
def new_shipment():
    with get_db() as db:
        departments = db.execute('SELECT * FROM department').fetchall()
        carrier_companies = db.execute('SELECT * FROM carrier_company').fetchall()
        shipment_types = db.execute('SELECT * FROM shipment_type').fetchall()
        governorates = db.execute('SELECT * FROM governorate').fetchall()
        
        if request.method == 'POST':
            # Validate required fields
            required_fields = [
                'shopiny_number', 'receipt_number', 'order_number', 'delivery_date',
                'source_governorate', 'destination_governorate', 'carrier_company'
            ]
            missing = [f for f in required_fields if not request.form.get(f)]
            # At least one item required
            if not any(k.startswith('items[0]') for k in request.form):
                missing.append('items')
            # Check items fields
            i = 0
            while f'items[{i}][shipment_type_id]' in request.form:
                for item_field in ['shipment_type_id', 'department_id', 'quantity', 'cost', 'boxes_count']:
                    if not request.form.get(f'items[{i}][{item_field}]'):
                        missing.append(f'items[{i}][{item_field}]')
                i += 1
            if missing:
                flash('جميع الحقول مطلوبة عدا الملاحظات. يرجى تعبئة جميع الحقول.', 'error')
                # إعادة عرض الصفحة مع نفس البيانات المدخلة وعدم تصفيرها
                return render_template('new_shipment.html',
                    departments=departments,
                    carrier_companies=carrier_companies,
                    shipment_types=shipment_types,
                    governorates=governorates,
                    form=request.form,
                    missing=missing)
            # تحقق من تكرار رقم الشحنة
            existing = db.execute('SELECT id FROM shipment WHERE shopiny_number = ?', [request.form['shopiny_number']]).fetchone()
            if existing:
                flash('رقم الشحنة مستخدم من قبل، يرجى اختيار رقم آخر.', 'error')
                return render_template('new_shipment.html',
                    departments=departments,
                    carrier_companies=carrier_companies,
                    shipment_types=shipment_types,
                    governorates=governorates,
                    form=request.form,
                    missing=missing)
            # تحقق من تكرار رقم وصل شحن جبال
            existing_receipt = db.execute('SELECT id FROM shipment WHERE receipt_number = ?', [request.form['receipt_number']]).fetchone()
            if existing_receipt:
                flash('رقم وصل شحن جبال مستخدم من قبل، يرجى اختيار رقم آخر.', 'error')
                return render_template('new_shipment.html',
                    departments=departments,
                    carrier_companies=carrier_companies,
                    shipment_types=shipment_types,
                    governorates=governorates,
                    form=request.form,
                    missing=missing)
            # First insert the shipment
            cursor = db.execute('''
                INSERT INTO shipment (
                    shopiny_number, receipt_number, order_number,
                    delivery_date, from_governorate, to_governorate,
                    carrier_company, notes
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', [
                request.form['shopiny_number'],
                request.form['receipt_number'],
                request.form['order_number'],
                datetime.strptime(request.form['delivery_date'], '%Y-%m-%d').strftime('%Y-%m-%d %H:%M:%S'),
                request.form['source_governorate'],
                request.form['destination_governorate'],
                request.form['carrier_company'],
                request.form.get('notes', '')
            ])
            
            shipment_id = cursor.lastrowid
            
            # Handle multiple items (total always boxes_count * cost)
            i = 0
            while f'items[{i}][shipment_type_id]' in request.form:
                quantity = int(request.form[f'items[{i}][quantity]'])
                cost = int(request.form[f'items[{i}][cost]'])
                boxes_count = int(request.form[f'items[{i}][boxes_count]'])
                use_boxes = 1 if request.form.get(f'items[{i}][use_boxes]') else 0
                # تعديل حساب total ليعتمد على use_boxes
                total = boxes_count * cost if use_boxes else quantity * cost
                db.execute('''
                    INSERT INTO shipment_item (
                        shipment_id, shipment_type_id, department_id,
                        quantity, cost, boxes_count, total, notes, use_boxes
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', [
                    shipment_id,
                    request.form[f'items[{i}][shipment_type_id]'],
                    request.form[f'items[{i}][department_id]'],
                    quantity,
                    cost,
                    boxes_count,
                    total,
                    request.form.get(f'items[{i}][notes]', ''),
                    use_boxes
                ])
                i += 1
                
            flash('تم إضافة الشحنة بنجاح', 'success')
            return redirect(url_for('index'))
            
    # توليد رقم وصل شحن جبال تلقائيًا عند GET فقط، أو عند POST إذا لم يكن موجودًا في form
    if request.method == 'POST':
        # إذا كان هناك بيانات مدخلة في الفورم (حتى لو كانت غير كاملة)، استخدمها كما هي
        shopiny_number = request.form.get('receipt_number', '')
    else:
        shopiny_number = ''
    return render_template('new_shipment.html',
                         departments=departments,
                         carrier_companies=carrier_companies,
                         shipment_types=shipment_types,
                         governorates=governorates,
                         form=request.form if request.method == 'POST' else None,
                         shopiny_number=shopiny_number)

@app.route('/shipment/<int:id>')
def view_shipment(id):
    with get_db() as db:
        shipment_row = db.execute('''
            SELECT *, datetime(delivery_date) as delivery_date 
            FROM shipment WHERE id = ?
        ''', [id]).fetchone()
        
        # Convert to dict and process date
        shipment = dict(shipment_row)
        try:
            shipment['delivery_date'] = datetime.strptime(shipment['delivery_date'], '%Y-%m-%d %H:%M:%S')
        except (ValueError, TypeError):
            shipment['delivery_date'] = None
        
        # Get items separately
        items = db.execute('''
            SELECT si.*, st.name as type_name, d.name as dept_name
            FROM shipment_item si
            JOIN shipment_type st ON si.shipment_type_id = st.id
            JOIN department d ON si.department_id = d.id
            WHERE si.shipment_id = ?
        ''', [id]).fetchall()
        # حساب المجموع الكلي
        total_sum = 0
        for item in items:
            try:
                # إذا كان الحقل use_boxes موجودًا ويعتمد عليه
                if 'use_boxes' in item.keys() and item['use_boxes']:
                    total_sum += int(item['boxes_count']) * int(item['cost'])
                else:
                    total_sum += int(item['quantity']) * int(item['cost'])
            except Exception:
                pass
    return render_template('view_shipment.html', shipment=shipment, items=items, total_sum=total_sum)

@app.route('/shipment/<int:id>/edit', methods=['GET', 'POST'])
def edit_shipment(id):
    with get_db() as db:
        # Fetch the shipment
        shipment_row = db.execute('''
            SELECT *, datetime(delivery_date) as delivery_date
            FROM shipment 
            WHERE id = ?
        ''', [id]).fetchone()
        
        # Convert to dict and process date
        shipment = dict(shipment_row)
        try:
            if shipment['delivery_date']:
                shipment['delivery_date'] = datetime.strptime(shipment['delivery_date'], '%Y-%m-%d %H:%M:%S')
        except (ValueError, TypeError):
            shipment['delivery_date'] = None
        
        # Fetch the items for this shipment with their related data
        items = db.execute('''
            SELECT si.*, st.name as type_name, d.name as dept_name
            FROM shipment_item si
            JOIN shipment_type st ON si.shipment_type_id = st.id
            JOIN department d ON si.department_id = d.id
            WHERE si.shipment_id = ?
        ''', [id]).fetchall()
        
        # Fetch lookup data for dropdowns
        departments = db.execute('SELECT * FROM department').fetchall()
        carrier_companies = db.execute('SELECT * FROM carrier_company').fetchall()
        shipment_types = db.execute('SELECT * FROM shipment_type').fetchall()
        governorates = db.execute('SELECT * FROM governorate').fetchall()
        
        if request.method == 'POST':
            # Validate required fields
            required_fields = [
                'shopiny_number', 'receipt_number', 'order_number', 'delivery_date',
                'source_governorate', 'destination_governorate', 'carrier_company'
            ]
            missing = [f for f in required_fields if not request.form.get(f)]
            # At least one item required
            if not any(k.startswith('items[0]') for k in request.form):
                missing.append('items')
            # Check items fields
            i = 0
            while f'items[{i}][shipment_type_id]' in request.form:
                for item_field in ['shipment_type_id', 'department_id', 'quantity', 'cost', 'boxes_count']:
                    if not request.form.get(f'items[{i}][{item_field}]'):
                        missing.append(f'items[{i}][{item_field}]')
                i += 1
            if missing:
                flash('جميع الحقول مطلوبة عدا الملاحظات. يرجى تعبئة جميع الحقول.', 'error')
                return render_template('edit_shipment.html',
                         shipment=shipment,
                         items=items,
                         departments=departments,
                         carrier_companies=carrier_companies,
                         shipment_types=shipment_types,
                         governorates=governorates,
                         form=request.form if request.method == 'POST' else None)
            # Update shipment
            db.execute('''
                UPDATE shipment SET
                    shopiny_number = ?, receipt_number = ?, 
                    order_number = ?, delivery_date = ?,
                    from_governorate = ?, to_governorate = ?,
                    carrier_company = ?, notes = ?
                WHERE id = ?
            ''', [
                request.form['shopiny_number'],
                request.form['receipt_number'],
                request.form['order_number'],
                datetime.strptime(request.form['delivery_date'], '%Y-%m-%d').strftime('%Y-%m-%d %H:%M:%S'),
                request.form['source_governorate'],
                request.form['destination_governorate'],
                request.form['carrier_company'],
                request.form.get('notes', ''),
                id
            ])
            
            db.execute('DELETE FROM shipment_item WHERE shipment_id = ?', [id])
            
            # Insert updated items (total always boxes_count * cost)
            i = 0
            while f'items[{i}][shipment_type_id]' in request.form:
                quantity = int(float(request.form[f'items[{i}][quantity]']))
                cost = int(float(request.form[f'items[{i}][cost]']))
                boxes_count = int(float(request.form[f'items[{i}][boxes_count]']))
                use_boxes = 1 if request.form.get(f'items[{i}][use_boxes]') else 0
                total = boxes_count * cost if use_boxes else quantity * cost
                db.execute('''
                    INSERT INTO shipment_item (
                        shipment_id, shipment_type_id, department_id,
                        quantity, cost, boxes_count, total, notes, use_boxes
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', [
                    id,
                    request.form[f'items[{i}][shipment_type_id]'],
                    request.form[f'items[{i}][department_id]'],
                    quantity,
                    cost,
                    boxes_count,
                    total,
                    request.form.get(f'items[{i}][notes]', ''),
                    use_boxes
                ])
                i += 1
            
            flash('تم تحديث الشحنة بنجاح', 'success')
            return redirect(url_for('index'))
            
    return render_template('edit_shipment.html',
                         shipment=shipment,
                         items=items,
                         departments=departments,
                         carrier_companies=carrier_companies,
                         shipment_types=shipment_types,
                         governorates=governorates,
                         form=request.form if request.method == 'POST' else None)

@app.route('/shipment/<int:id>/delete', methods=['POST'])
def delete_shipment(id):
    with get_db() as db:
        db.execute('DELETE FROM shipment_item WHERE shipment_id = ?', [id])
        db.execute('DELETE FROM shipment WHERE id = ?', [id])
        flash('تم حذف الشحنة بنجاح', 'success')
    return redirect(url_for('index'))

@app.route('/shipment-type/new', methods=['GET', 'POST'])
def new_shipment_type():
    if request.method == 'POST':
        with get_db() as db:
            db.execute('''
                INSERT INTO shipment_type (name) VALUES (?)
            ''', [request.form['name']])
            flash('تم إضافة نوع الشحنة بنجاح', 'success')
            return redirect(url_for('index'))
    return render_template('new_shipment_type.html')

@app.route('/department/new', methods=['GET', 'POST'])
def new_department():
    if request.method == 'POST':
        with get_db() as db:
            db.execute('''
                INSERT INTO department (name) VALUES (?)
            ''', [request.form['name']])
            flash('تم إضافة القسم بنجاح', 'success')
            return redirect(url_for('index'))
    return render_template('new_department.html')

@app.route('/shipment-types')
def list_shipment_types():
    with get_db() as db:
        shipment_types = db.execute('SELECT * FROM shipment_type').fetchall()
    return render_template('list_shipment_types.html', shipment_types=shipment_types)

@app.route('/shipment-type/<int:id>/edit', methods=['GET', 'POST'])
def edit_shipment_type(id):
    with get_db() as db:
        shipment_type = db.execute('SELECT * FROM shipment_type WHERE id = ?', [id]).fetchone()
        if request.method == 'POST':
            db.execute('''
                UPDATE shipment_type SET name = ? WHERE id = ?
            ''', [request.form['name'], id])
            flash('تم تحديث نوع الشحنة بنجاح', 'success')
            return redirect(url_for('list_shipment_types'))
    return render_template('edit_shipment_type.html', shipment_type=shipment_type)

@app.route('/shipment-type/<int:id>/delete', methods=['POST'])
def delete_shipment_type(id):
    with get_db() as db:
        db.execute('DELETE FROM shipment_type WHERE id = ?', [id])
        flash('تم حذف نوع الشحنة بنجاح', 'success')
    return redirect(url_for('list_shipment_types'))

@app.route('/departments')
def list_departments():
    with get_db() as db:
        departments = db.execute('SELECT * FROM department').fetchall()
    return render_template('list_departments.html', departments=departments)

@app.route('/department/<int:id>/edit', methods=['GET', 'POST'])
def edit_department(id):
    with get_db() as db:
        department = db.execute('SELECT * FROM department WHERE id = ?', [id]).fetchone()
        if request.method == 'POST':
            db.execute('''
                UPDATE department SET name = ? WHERE id = ?
            ''', [request.form['name'], id])
            flash('تم تحديث القسم بنجاح', 'success')
            return redirect(url_for('list_departments'))
    return render_template('edit_department.html', department=department)

@app.route('/department/<int:id>/delete', methods=['POST'])
def delete_department(id):
    with get_db() as db:
        db.execute('DELETE FROM department WHERE id = ?', [id])
        flash('تم حذف القسم بنجاح', 'success')
    return redirect(url_for('list_departments'))

@app.route('/carrier-companies')
def list_carrier_companies():
    with get_db() as db:
        companies = db.execute('SELECT * FROM carrier_company').fetchall()
    return render_template('list_carrier_companies.html', companies=companies)

@app.route('/carrier-company/new', methods=['GET', 'POST'])
def new_carrier_company():
    if request.method == 'POST':
        with get_db() as db:
            db.execute('''
                INSERT INTO carrier_company (name) VALUES (?)
            ''', [request.form['name']])
            flash('تم إضافة شركة النقل بنجاح', 'success')
            return redirect(url_for('list_carrier_companies'))
    return render_template('new_carrier_company.html')

@app.route('/carrier-company/<int:id>/edit', methods=['GET', 'POST'])
def edit_carrier_company(id):
    with get_db() as db:
        company = db.execute('SELECT * FROM carrier_company WHERE id = ?', [id]).fetchone()
        if request.method == 'POST':
            db.execute('''
                UPDATE carrier_company SET name = ? WHERE id = ?
            ''', [request.form['name'], id])
            flash('تم تحديث شركة النقل بنجاح', 'success')
            return redirect(url_for('list_carrier_companies'))
    return render_template('edit_carrier_company.html', company=company)

@app.route('/carrier-company/<int:id>/delete', methods=['POST'])
def delete_carrier_company(id):
    with get_db() as db:
        db.execute('DELETE FROM carrier_company WHERE id = ?', [id])
        flash('تم حذف شركة النقل بنجاح', 'success')
    return redirect(url_for('list_carrier_companies'))

@app.route('/governorates')
def list_governorates():
    with get_db() as db:
        governorates = db.execute('SELECT * FROM governorate').fetchall()
    return render_template('list_governorates.html', governorates=governorates)

@app.route('/governorate/new', methods=['GET', 'POST'])
def new_governorate():
    if request.method == 'POST':
        with get_db() as db:
            db.execute('INSERT INTO governorate (name) VALUES (?)',
                      [request.form['name']])
            flash('تم إضافة المحافظة بنجاح', 'success')
            return redirect(url_for('list_governorates'))
    return render_template('new_governorate.html')

@app.route('/governorate/<int:id>/edit', methods=['GET', 'POST'])
def edit_governorate(id):
    with get_db() as db:
        governorate = db.execute('SELECT * FROM governorate WHERE id = ?',
                               [id]).fetchone()
        if request.method == 'POST':
            db.execute('UPDATE governorate SET name = ? WHERE id = ?',
                      [request.form['name'], id])
            flash('تم تحديث المحافظة بنجاح', 'success')
            return redirect(url_for('list_governorates'))
    return render_template('edit_governorate.html', governorate=governorate)

@app.route('/governorate/<int:id>/delete', methods=['POST'])
def delete_governorate(id):
    with get_db() as db:
        db.execute('DELETE FROM governorate WHERE id = ?', [id])
        flash('تم حذف المحافظة بنجاح', 'success')
    return redirect(url_for('list_governorates'))

@app.route('/reports/monthly', methods=['GET', 'POST'])
def reports_monthly():
    from_date = request.form.get('from_date')
    to_date = request.form.get('to_date')
    shipments = []
    items = []
    with get_db() as db:
        # تعديل الاستعلام ليكون بدون ربط shipment_item وضمان عدم تكرار الشحنات
        # تعديل الترتيب ليكون تصاعديًا حسب تاريخ تسليم الشحنة
        query = '''
            SELECT DISTINCT s.id, s.*, c.name as carrier_company_name, g1.name as from_gov_name, g2.name as to_gov_name
            FROM shipment s
            LEFT JOIN carrier_company c ON s.carrier_company = c.name
            LEFT JOIN governorate g1 ON s.from_governorate = g1.name
            LEFT JOIN governorate g2 ON s.to_governorate = g2.name
            WHERE 1=1
        '''
        params = []
        if from_date:
            query += ' AND date(s.delivery_date) >= ?'
            params.append(from_date)
        if to_date:
            query += ' AND date(s.delivery_date) <= ?'
            params.append(to_date)
        # تعديل ترتيب النتائج ليكون تصاعديًا حسب تاريخ التسليم
        query += ' GROUP BY s.id ORDER BY s.delivery_date ASC'
        shipments = db.execute(query, params).fetchall()
        
        # جلب البنود بشكل منفصل
        shipment_ids = [str(s['id']) for s in shipments]
        items = []
        if shipment_ids:
            q = f"SELECT si.*, st.name as shipment_type_name, d.name as department_name FROM shipment_item si JOIN shipment_type st ON si.shipment_type_id=st.id JOIN department d ON si.department_id=d.id WHERE si.shipment_id IN ({','.join(['?']*len(shipment_ids))})"
            items = db.execute(q, shipment_ids).fetchall()
    
    # تنظيم البنود حسب رقم الشحنة
    items_by_shipment = {}
    total_cost = 0
    total_quantity = 0
    total_boxes = 0
    
    for item in items:
        items_by_shipment.setdefault(item['shipment_id'], []).append(item)
        # تعديل حساب مجموع كلفة الشحنة ليعتمد على use_boxes
        try:
            quantity = int(item['quantity'])
            cost = int(item['cost'])
            boxes_count = int(item['boxes_count'])
            use_boxes = bool(item['use_boxes']) if 'use_boxes' in item.keys() else False
            
            # حساب كلفة الشحنة بناءً على قيمة use_boxes
            item_cost = (boxes_count if use_boxes else quantity) * cost
            
            # إضافة إلى المجاميع
            total_quantity += quantity
            total_boxes += boxes_count
            total_cost += item_cost
        except Exception:
            pass
    
    # تحويل Shipments إلى قائمة عادية بدل sqlite3.Row لمنع التكرار المحتمل
    shipments_list = []
    seen_ids = set()
    for s in shipments:
        if s['id'] not in seen_ids:
            seen_ids.add(s['id'])
            shipments_list.append(dict(s))
    
    return render_template('reports_monthly.html', 
                         shipments=shipments_list, 
                         items_by_shipment=items_by_shipment, 
                         from_date=from_date, 
                         to_date=to_date, 
                         total_cost=total_cost,
                         total_quantity=total_quantity,
                         total_boxes=total_boxes)

@app.route('/reports/monthly/export', methods=['POST'])
def export_monthly_report():
    from_date = request.form.get('from_date')
    to_date = request.form.get('to_date')
    with get_db() as db:
        # تعديل الاستعلام ليكون بدون ربط shipment_item وضمان عدم تكرار الشحنات
        # تعديل الترتيب ليكون تصاعديًا حسب تاريخ تسليم الشحنة
        query = '''
            SELECT DISTINCT s.id, s.*, c.name as carrier_company_name, g1.name as from_gov_name, g2.name as to_gov_name
            FROM shipment s
            LEFT JOIN carrier_company c ON s.carrier_company = c.name
            LEFT JOIN governorate g1 ON s.from_governorate = g1.name
            LEFT JOIN governorate g2 ON s.to_governorate = g2.name
            WHERE 1=1
        '''
        params = []
        if from_date:
            query += ' AND date(s.delivery_date) >= ?'
            params.append(from_date)
        if to_date:
            query += ' AND date(s.delivery_date) <= ?'
            params.append(to_date)
        # تعديل ترتيب النتائج ليكون تصاعديًا حسب تاريخ التسليم
        query += ' GROUP BY s.id ORDER BY s.delivery_date ASC'
        shipments = db.execute(query, params).fetchall()
        
        # جلب البنود بشكل منفصل
        shipment_ids = [str(s['id']) for s in shipments]
        items = []
        if shipment_ids:
            q = f"SELECT si.*, st.name as shipment_type_name, d.name as department_name FROM shipment_item si JOIN shipment_type st ON si.shipment_type_id=st.id JOIN department d ON si.department_id=d.id WHERE si.shipment_id IN ({','.join(['?']*len(shipment_ids))})"
            items = db.execute(q, shipment_ids).fetchall()
    
    # تنظيم البنود حسب رقم الشحنة
    items_by_shipment = {}
    for item in items:
        items_by_shipment.setdefault(item['shipment_id'], []).append(item)
    
    # حساب المجاميع
    total_sum = 0
    total_quantity = 0
    total_boxes = 0
    
    for item in items:
        try:
            quantity = int(item['quantity'])
            cost = int(item['cost'])
            boxes_count = int(item['boxes_count'])
            use_boxes = bool(item['use_boxes']) if 'use_boxes' in item.keys() else False
            
            # حساب كلفة الشحنة بنفس طريقة النموذج
            calculated_total = (boxes_count if use_boxes else quantity) * cost
            
            # إضافة إلى المجاميع
            total_quantity += quantity
            total_boxes += boxes_count
            total_sum += calculated_total  # استخدام القيمة المحسوبة بدلاً من total المخزنة
        except Exception:
            pass
    
    # إنشاء ملف الإكسل
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet('تقرير الشحنات الشهري')
    worksheet.right_to_left()  # عرض الجدول من اليمين إلى اليسار
    # إعداد التنسيقات
    title_format = workbook.add_format({'bold': True, 'font_size': 18, 'align': 'center', 'valign': 'vcenter', 'font_color': '#1976d2'})
    date_format = workbook.add_format({'font_size': 12, 'align': 'center', 'valign': 'vcenter', 'font_color': '#333'})
    header_format = workbook.add_format({'bold': True, 'bg_color': '#2196F3', 'font_color': 'white', 'border': 1, 'align': 'center', 'valign': 'vcenter'})
    details_header_format = workbook.add_format({'bold': True, 'bg_color': '#ffe0b2', 'font_color': '#b26a00', 'border': 1, 'align': 'center', 'valign': 'vcenter'})
    cell_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1})
    total_format = workbook.add_format({'bold': True, 'bg_color': '#ffe082', 'font_color': '#b26a00', 'align': 'center', 'valign': 'vcenter', 'border': 1, 'font_size': 14})
    total_value_format = workbook.add_format({'bold': True, 'bg_color': '#fff8e1', 'font_color': '#b26a00', 'align': 'center', 'valign': 'vcenter', 'border': 1, 'font_size': 14})
    # عنوان التقرير
    worksheet.merge_range(0, 0, 0, 14, 'تقرير الشحنات الشهري', title_format)
    # تاريخ من وإلى
    date_range = f"من: {from_date or '-'}   إلى: {to_date or '-'}"
    worksheet.merge_range(1, 0, 1, 14, date_range, date_format)
    # تعديل عناوين الأعمدة لتطابق النموذج
    details_headers = ['صنف الشحنة', 'القسم', 'الكمية', 'التكلفة', 'عدد الكراتين', 'كلفة الشحنة', 'ملاحظات الشحنة']
    # رؤوس الأعمدة
    headers = [
        '#', 'رقم شحنة شركة النقل', 'تاريخ تسليم الشحنة', 'رقم وصل شحن جبال', 'رقم الأوردر',
        'من', 'إلى', 'الشركة الناقلة',
        'تفاصيل الشحنة'
    ]
    # دمج الخلايا لرأس تفاصيل الشحنة فقط بعدد أعمدة التفاصيل (من 8 إلى 14)
    worksheet.merge_range(2, 8, 2, 14, 'تفاصيل الشحنة', details_header_format)
    # كتابة رؤوس الأعمدة الأخرى
    for col, h in enumerate(headers):
        if col < 8:
            worksheet.write(2, col, h, header_format)
    
    # كتابة عناوين تفاصيل الشحنة
    for col, h in enumerate(details_headers):
        worksheet.write(3, col + 8, h, details_header_format)
    
    # تعديل بداية الصفوف للبيانات
    row = 4
    row_num = 1
    for s in shipments:
        shipment_items = items_by_shipment.get(s['id'], [])
        rowspan = len(shipment_items) if shipment_items else 1
        if shipment_items:
            # دمج خلايا الأعمدة الأساسية إذا كان هناك أكثر من بند
            if rowspan > 1:
                for col, value in enumerate([
                    row_num, s['shopiny_number'], str(s['delivery_date'])[:10] if s['delivery_date'] else '',
                    s['receipt_number'], s['order_number'], s['from_gov_name'], s['to_gov_name'], s['carrier_company_name']
                ]):
                    worksheet.merge_range(row, col, row + rowspan - 1, col, value, cell_format)
            else:
                for col, value in enumerate([
                    row_num, s['shopiny_number'], str(s['delivery_date'])[:10] if s['delivery_date'] else '',
                    s['receipt_number'], s['order_number'], s['from_gov_name'], s['to_gov_name'], s['carrier_company_name']
                ]):
                    worksheet.write(row, col, value, cell_format)
            
            for idx, item in enumerate(shipment_items):
                item_row = row + idx
                
                # قراءة القيم مباشرة من البيانات
                quantity = int(item['quantity'])
                cost = int(item['cost'])
                boxes_count = int(item['boxes_count'])
                use_boxes = bool(item['use_boxes']) if 'use_boxes' in item.keys() else False
                
                # حساب كلفة الشحنة بنفس طريقة النموذج
                calculated_total = (boxes_count if use_boxes else quantity) * cost
                
                # كتابة البيانات في الخلايا
                worksheet.write(item_row, 8, item['shipment_type_name'], cell_format)  # صنف الشحنة
                worksheet.write(item_row, 9, item['department_name'], cell_format)     # القسم
                worksheet.write(item_row, 10, quantity, cell_format)                   # الكمية
                worksheet.write(item_row, 11, cost, cell_format)                       # التكلفة
                worksheet.write(item_row, 12, boxes_count, cell_format)                # عدد الكراتين
                worksheet.write(item_row, 13, '{:,}'.format(calculated_total), cell_format)       # كلفة الشحنة
                worksheet.write(item_row, 14, item['notes'], cell_format)              # ملاحظات الشحنة
            
            row += rowspan
            row_num += 1
        else:
            worksheet.write(row, 0, row_num, cell_format)
            worksheet.write(row, 1, s['shopiny_number'], cell_format)
            worksheet.write(row, 2, str(s['delivery_date'])[:10] if s['delivery_date'] else '', cell_format)
            worksheet.write(row, 3, s['receipt_number'], cell_format)
            worksheet.write(row, 4, s['order_number'], cell_format)
            worksheet.write(row, 5, s['from_gov_name'], cell_format)
            worksheet.write(row, 6, s['to_gov_name'], cell_format)
            worksheet.write(row, 7, s['carrier_company_name'], cell_format)
            worksheet.merge_range(row, 8, row, 14, 'لا يوجد بنود', cell_format)
            row += 1
            row_num += 1
    
    # صف المجموع الكلي (موحد)
    worksheet.merge_range(row, 0, row, 9, 'المجموع الكلي:', total_format)
    worksheet.write(row, 10, '{:,}'.format(total_quantity), total_value_format)  # الكمية
    worksheet.write(row, 11, '', total_value_format)  # التكلفة (فارغة)
    worksheet.write(row, 12, '{:,}'.format(total_boxes), total_value_format)  # عدد الكراتين
    worksheet.write(row, 13, '{:,}'.format(total_sum), total_value_format)  # كلفة الشحنة
    worksheet.write(row, 14, '', total_value_format)
    
    # ضبط عرض الأعمدة
    worksheet.set_column(0, 0, 5)
    worksheet.set_column(1, 1, 18)
    worksheet.set_column(2, 2, 15)
    worksheet.set_column(3, 4, 15)
    worksheet.set_column(5, 7, 13)
    worksheet.set_column(8, 14, 13)
    
    workbook.close()
    output.seek(0)
    return send_file(output, download_name='monthly_report.xlsx', as_attachment=True)

@app.route('/reports/by', methods=['GET', 'POST'])
def reports_by():
    from_date = request.form.get('from_date')
    to_date = request.form.get('to_date')
    carrier_company = request.form.get('carrier_company')
    shipment_type = request.form.get('shipment_type')
    department = request.form.get('department')
    shipments = []
    items = []
    with get_db() as db:
        # جلب القوائم المنسدلة
        carrier_companies = db.execute('SELECT name FROM carrier_company').fetchall()
        shipment_types = db.execute('SELECT id, name FROM shipment_type').fetchall()
        departments = db.execute('SELECT id, name FROM department').fetchall()
        query = '''
            SELECT s.*, c.name as carrier_company_name, g1.name as from_gov_name, g2.name as to_gov_name
            FROM shipment s
            LEFT JOIN carrier_company c ON s.carrier_company = c.name
            LEFT JOIN governorate g1 ON s.from_governorate = g1.name
            LEFT JOIN governorate g2 ON s.to_governorate = g2.name
            WHERE 1=1
        '''
        params = []
        if from_date:
            query += ' AND date(s.delivery_date) >= ?'
            params.append(from_date)
        if to_date:
            query += ' AND date(s.delivery_date) <= ?'
            params.append(to_date)
        if carrier_company:
            query += ' AND s.carrier_company = ?'
            params.append(carrier_company)
        query += ' ORDER BY s.delivery_date ASC'  # تعديل الترتيب ليكون تصاعديًا حسب التاريخ
        shipments = db.execute(query, params).fetchall()
        shipment_ids = [str(s['id']) for s in shipments]
        items = []
        if shipment_ids:
            item_query = '''SELECT si.*, st.name as shipment_type_name, d.name as department_name FROM shipment_item si \
                JOIN shipment_type st ON si.shipment_type_id=st.id \
                JOIN department d ON si.department_id=d.id \
                WHERE si.shipment_id IN ({})'''.format(','.join(['?']*len(shipment_ids)))
            item_params = shipment_ids
            # فلترة البنود حسب الصنف أو القسم
            if shipment_type:
                item_query += ' AND si.shipment_type_id = ?'
                item_params.append(shipment_type)
            if department:
                item_query += ' AND si.department_id = ?'
                item_params.append(department)
            items = db.execute(item_query, item_params).fetchall()
    
    items_by_shipment = {}
    total_cost = 0
    total_quantity = 0
    total_boxes = 0
    
    for item in items:
        items_by_shipment.setdefault(item['shipment_id'], []).append(item)
        try:
            quantity = int(item['quantity'])
            cost = int(item['cost'])
            boxes_count = int(item['boxes_count'])
            use_boxes = bool(item['use_boxes']) if 'use_boxes' in item.keys() else False
            
            # حساب كلفة الشحنة بناءً على قيمة use_boxes
            item_cost = (boxes_count if use_boxes else quantity) * cost
            
            # إضافة إلى المجاميع
            total_quantity += quantity
            total_boxes += boxes_count
            total_cost += item_cost
        except Exception:
            pass
    
    return render_template('reports_by.html', 
                         shipments=shipments, 
                         items_by_shipment=items_by_shipment, 
                         from_date=from_date, 
                         to_date=to_date, 
                         total_cost=total_cost,
                         total_quantity=total_quantity,
                         total_boxes=total_boxes,
                         carrier_companies=carrier_companies, 
                         shipment_types=shipment_types, 
                         departments=departments, 
                         selected_carrier=carrier_company, 
                         selected_type=shipment_type, 
                         selected_dept=department)

@app.route('/reports/by/export', methods=['POST'])
def export_by_report():
    from_date = request.form.get('from_date')
    to_date = request.form.get('to_date')
    carrier_company = request.form.get('carrier_company')
    shipment_type = request.form.get('shipment_type')
    department = request.form.get('department')
    with get_db() as db:
        query = '''
            SELECT s.*, c.name as carrier_company_name, g1.name as from_gov_name, g2.name as to_gov_name
            FROM shipment s
            LEFT JOIN carrier_company c ON s.carrier_company = c.name
            LEFT JOIN governorate g1 ON s.from_governorate = g1.name
            LEFT JOIN governorate g2 ON s.to_governorate = g2.name
            WHERE 1=1
        '''
        params = []
        if from_date:
            query += ' AND date(s.delivery_date) >= ?'
            params.append(from_date)
        if to_date:
            query += ' AND date(s.delivery_date) <= ?'
            params.append(to_date)
        if carrier_company:
            query += ' AND s.carrier_company = ?'
            params.append(carrier_company)
        query += ' ORDER BY s.delivery_date ASC'
        shipments = db.execute(query, params).fetchall()
        
        # استعلام للحصول على البنود التي تتطابق مع الفلتر
        shipment_ids = [str(s['id']) for s in shipments]
        items = []
        if shipment_ids:
            item_query = '''SELECT si.*, st.name as shipment_type_name, d.name as department_name FROM shipment_item si \
                JOIN shipment_type st ON si.shipment_type_id=st.id \
                JOIN department d ON si.department_id=d.id \
                WHERE si.shipment_id IN ({})'''.format(','.join(['?']*len(shipment_ids)))
            item_params = shipment_ids
            if shipment_type:
                item_query += ' AND si.shipment_type_id = ?'
                item_params.append(shipment_type)
            if department:
                item_query += ' AND si.department_id = ?'
                item_params.append(department)
            items = db.execute(item_query, item_params).fetchall()
    
    # تنظيم البنود حسب رقم الشحنة
    items_by_shipment = {}
    total_sum = 0
    total_quantity = 0
    total_boxes = 0
    
    for item in items:
        items_by_shipment.setdefault(item['shipment_id'], []).append(item)
        # حساب القيم الكلية باستخدام حقل total المخزن مباشرة
        try:
            quantity = int(item['quantity'])
            boxes_count = int(item['boxes_count'])
            total = int(item['total'])
            
            # إضافة إلى المجاميع
            total_quantity += quantity
            total_boxes += boxes_count
            total_sum += total  # استخدام قيمة total المخزنة مباشرة
        except Exception:
            pass
    
    # إزالة الشحنات التي لا تحتوي على بنود متطابقة مع الفلتر
    filtered_shipments = []
    for s in shipments:
        if s['id'] in items_by_shipment:
            filtered_shipments.append(s)
    
    # إنشاء ملف الإكسل
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet('تقرير الشحنات')
    worksheet.right_to_left()
    
    # إعداد التنسيقات
    title_format = workbook.add_format({'bold': True, 'font_size': 18, 'align': 'center', 'valign': 'vcenter', 'font_color': '#1976d2'})
    date_format = workbook.add_format({'font_size': 12, 'align': 'center', 'valign': 'vcenter', 'font_color': '#333'})
    header_format = workbook.add_format({'bold': True, 'bg_color': '#2196F3', 'font_color': 'white', 'border': 1, 'align': 'center', 'valign': 'vcenter'})
    details_header_format = workbook.add_format({'bold': True, 'bg_color': '#ffe0b2', 'font_color': '#b26a00', 'border': 1, 'align': 'center', 'valign': 'vcenter'})
    cell_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1})
    total_format = workbook.add_format({'bold': True, 'bg_color': '#ffe082', 'font_color': '#b26a00', 'align': 'center', 'valign': 'vcenter', 'border': 1, 'font_size': 14})
    total_value_format = workbook.add_format({'bold': True, 'bg_color': '#fff8e1', 'font_color': '#b26a00', 'align': 'center', 'valign': 'vcenter', 'border': 1, 'font_size': 14})
    
    # عنوان التقرير
    worksheet.merge_range(0, 0, 0, 14, 'تقرير الشحنات', title_format)
    
    # تاريخ من وإلى
    date_range = f"من: {from_date or '-'}   إلى: {to_date or '-'}"
    worksheet.merge_range(1, 0, 1, 14, date_range, date_format)
    
    # تعديل عناوين الأعمدة لتطابق النموذج
    details_headers = ['صنف الشحنة', 'الكمية', 'التكلفة', 'عدد الصناديق', 'الإجمالي', 'ملاحظات الشحنة']
    
    # دمج الخلايا لرأس تفاصيل الشحنة فقط بعدد أعمدة التفاصيل (من 8 إلى 14)
    worksheet.merge_range(2, 8, 2, 14, 'تفاصيل الشحنة', details_header_format)
    
    # رؤوس الأعمدة الأساسية
    headers = [
        '#', 'رقم شحنة شركة النقل', 'تاريخ تسليم الشحنة', 'رقم وصل شحن جبال', 'رقم الأوردر',
        'من', 'إلى', 'الشركة الناقلة'
    ]
    
    # كتابة رؤوس الأعمدة الأساسية
    for col, h in enumerate(headers):
        worksheet.write(2, col, h, header_format)
    
    # كتابة عناوين تفاصيل الشحنة
    for col, h in enumerate(details_headers):
        worksheet.write(3, col + 8, h, details_header_format)
    
    row = 4
    row_num = 1
    
    for s in filtered_shipments:
        shipment_items = items_by_shipment.get(s['id'], [])
        rowspan = len(shipment_items)
        
        # دمج خلايا الأعمدة الأساسية إذا كان هناك أكثر من بند
        if rowspan > 1:
            for col, value in enumerate([
                row_num, s['shopiny_number'], str(s['delivery_date'])[:10] if s['delivery_date'] else '',
                s['receipt_number'], s['order_number'], s['from_gov_name'], s['to_gov_name'], s['carrier_company_name']
            ]):
                worksheet.merge_range(row, col, row + rowspan - 1, col, value, cell_format)
        else:
            for col, value in enumerate([
                row_num, s['shopiny_number'], str(s['delivery_date'])[:10] if s['delivery_date'] else '',
                s['receipt_number'], s['order_number'], s['from_gov_name'], s['to_gov_name'], s['carrier_company_name']
            ]):
                worksheet.write(row, col, value, cell_format)
        
        for idx, item in enumerate(shipment_items):
            item_row = row + idx
            
            # حساب قيم العرض
            quantity = int(item['quantity'])
            cost = int(item['cost'])
            boxes_count = int(item['boxes_count'])
            use_boxes = bool(item['use_boxes']) if 'use_boxes' in item.keys() else False
            
            # حساب كلفة الشحنة بنفس طريقة النموذج
            calculated_total = (boxes_count if use_boxes else quantity) * cost
            
            # كتابة البيانات في الخلايا
            worksheet.write(item_row, 8, item['shipment_type_name'], cell_format)  # صنف الشحنة
            worksheet.write(item_row, 9, item['department_name'], cell_format)     # القسم
            worksheet.write(item_row, 10, quantity, cell_format)                   # الكمية
            worksheet.write(item_row, 11, cost, cell_format)                       # التكلفة
            worksheet.write(item_row, 12, boxes_count, cell_format)                # عدد الكراتين
            worksheet.write(item_row, 13, '{:,}'.format(calculated_total), cell_format)       # كلفة الشحنة
            worksheet.write(item_row, 14, item['notes'], cell_format)              # ملاحظات الشحنة
        
        row += rowspan
        row_num += 1
    
    # صف المجموع الكلي (موحد)
    worksheet.merge_range(row, 0, row, 9, 'مجموع تكلفة النقل الكلية:', total_format)
    worksheet.write(row, 10, '{:,}'.format(total_quantity), total_value_format)  # الكمية
    worksheet.write(row, 11, '', total_value_format)  # التكلفة (فارغة)
    worksheet.write(row, 12, '{:,}'.format(total_boxes), total_value_format)  # عدد الكراتين
    worksheet.write(row, 13, '{:,}'.format(total_sum), total_value_format)  # كلفة الشحنة
    worksheet.write(row, 14, '', total_value_format)
    worksheet.set_column(0, 0, 5)
    worksheet.set_column(1, 1, 18)
    worksheet.set_column(2, 2, 15)
    worksheet.set_column(3, 4, 15)
    worksheet.set_column(5, 7, 13)
    worksheet.set_column(8, 14, 13)
    workbook.close()
    output.seek(0)
    return send_file(output, download_name='shipments_report.xlsx', as_attachment=True)

@app.route('/api/generate_shopiny_number')
def api_generate_shopiny_number():
    governorate = request.args.get('governorate')
    delivery_date = request.args.get('delivery_date')
    with get_db() as db:
        number = generate_shopiny_number(db, governorate, delivery_date)
    return {'shopiny_number': number}

def generate_shopiny_number(db, governorate, delivery_date):
    """
    توليد رقم وصل شحن جبال: اسم المحافظة + اخر رقمين من السنة + رقمين من الشهر + رقم تسلسلي 4 مراتب
    مع التأكد من عدم تكرار الرقم في قاعدة البيانات
    """
    if not governorate or not delivery_date:
        return ''
    # استخدم اسم المحافظة كما هو
    gov_part = str(governorate)
    # التاريخ
    try:
        dt = datetime.strptime(delivery_date, '%Y-%m-%d')
    except Exception:
        return ''
    year_part = str(dt.year)[-2:]
    month_part = f'{dt.month:02d}'
    # احسب الرقم التسلسلي لهذا الشهر والمحافظة بناءً على receipt_number
    serial_query = '''SELECT receipt_number FROM shipment WHERE from_governorate = ? AND strftime('%Y-%m', delivery_date) = ? AND receipt_number IS NOT NULL AND receipt_number != '' ORDER BY id DESC LIMIT 1'''
    last = db.execute(serial_query, [governorate, dt.strftime('%Y-%m')]).fetchone()
    if last and last['receipt_number']:
        try:
            last_serial = int(last['receipt_number'][-4:])
        except Exception:
            last_serial = 0
    else:
        last_serial = 0
    
    # البدء بالرقم التسلسلي التالي للآخر المستخدم
    new_serial = last_serial + 1
    
    # التأكد من أن الرقم المُنشأ غير موجود بالفعل في قاعدة البيانات
    while True:
        serial_part = f'{new_serial:04d}'
        generated_number = f'{gov_part}{year_part}{month_part}{serial_part}'
        
        # التحقق من وجود هذا الرقم في قاعدة البيانات
        existing = db.execute('SELECT id FROM shipment WHERE receipt_number = ?', [generated_number]).fetchone()
        if not existing:
            # الرقم غير موجود، يمكن استخدامه
            break
            
        # إذا كان الرقم موجودًا، جرب الرقم التالي
        new_serial += 1
        
        # منع الحلقة اللانهائية في حالة وجود مشاكل غير متوقعة
        if new_serial > last_serial + 1000:  # حد أقصى معقول
            break
    
    return generated_number

@app.route('/export-all')
def export_all():
    """
    Export all tables to a single Excel file (each table as a sheet).
    """
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    with get_db() as db:
        tables = [
            'shipment', 'shipment_item', 'shipment_type',
            'department', 'carrier_company', 'governorate'
        ]
        for table in tables:
            worksheet = workbook.add_worksheet(table)
            rows = db.execute(f"SELECT * FROM {table}").fetchall()
            if not rows:
                continue
            # Write headers
            headers = rows[0].keys()
            for col, h in enumerate(headers):
                worksheet.write(0, col, h)
            # Write data
            for row_idx, row in enumerate(rows, 1):
                for col, h in enumerate(headers):
                    worksheet.write(row_idx, col, row[h])
    workbook.close()
    output.seek(0)
    return send_file(output, download_name='all_data.xlsx', as_attachment=True)

@app.route('/import-all', methods=['GET', 'POST'])
def import_all():
    """
    Import all tables from an uploaded Excel file.
    """
    if openpyxl is None:
        flash('مكتبة openpyxl غير مثبتة. يرجى تثبيتها أولاً (pip install openpyxl).', 'error')
        return redirect(url_for('index'))
    if request.method == 'POST':
        file = request.files.get('import_file')
        if not file:
            flash('يرجى اختيار ملف إكسل.', 'error')
            return redirect(url_for('import_all'))
        wb = openpyxl.load_workbook(file)
        tables = [
            'shipment', 'shipment_item', 'shipment_type',
            'department', 'carrier_company', 'governorate'
        ]
        with get_db() as db:
            db.execute('PRAGMA foreign_keys = OFF')
            # Clear all tables
            for table in tables:
                db.execute(f'DELETE FROM {table}')
            # Import each sheet
            for table in tables:
                if table not in wb.sheetnames:
                    continue
                ws = wb[table]
                rows = list(ws.iter_rows(values_only=True))
                if not rows or len(rows) < 2:
                    continue
                headers = [str(h) for h in rows[0]]
                for data_row in rows[1:]:
                    # Skip empty rows
                    if all(cell is None for cell in data_row):
                        continue
                    placeholders = ','.join(['?'] * len(headers))
                    db.execute(
                        f'INSERT INTO {table} ({",".join(headers)}) VALUES ({placeholders})',
                        data_row
                    )
            db.execute('PRAGMA foreign_keys = ON')
        flash('تم استيراد جميع البيانات بنجاح.', 'success')
        return redirect(url_for('index'))
    return render_template('import_all.html')

# Initialize the database when the app starts
with app.app_context():
    init_db()

if __name__ == '__main__':
    app.run(debug=True)